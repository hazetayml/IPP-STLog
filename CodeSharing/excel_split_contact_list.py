import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook("Python Excel Student Data.xlsx")
sheet=workbook["Sheet1"]

# Get the unique ReqTerm values
req_terms = set()
for i in range(2, sheet.max_row + 1):
    req_term = sheet.cell(row=i, column=1).value
    req_terms.add(req_term)

# Create a new workbook
new_wb = openpyxl.Workbook()
# Delete the default sheet created when initializing a Workbook
new_wb.remove(new_wb.active)

# Create a sheet for each unique ReqTerm and populate it with the relevant rows
for term in req_terms:
    new_sheet = new_wb.create_sheet(title=str(term))
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == term or i == 1:
            values = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
            new_sheet.append(values)

# Save the new workbook
new_wb.save("output.xlsx")
